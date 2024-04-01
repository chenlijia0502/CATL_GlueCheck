import numpy as np
import cv2
import os

import openpyxl


class CExcelManager(object):
    """
    读写EXCEL的类
    """
    def __init__(self,file_name, sheet_name, head):
        """
        这个是用来初始化读取对象的
        :param file_name: 文件名 ---> str类型
        :param sheet_name: 表单名 ———> str类型
        """
        # 打开文件
        self.__createfile(file_name, sheet_name, head)
        self.wb = openpyxl.load_workbook(file_name)

        self.file_name = file_name
        # 选择表单
        self.sh = self.wb[sheet_name]
        self.b_iswriting = False# 是否正在写的状态


    def __createfile(self, file_name, sheet_name, head):
        """
        创建文件，包含创建文件夹
        file_name: 文件名
        sheet_name: 工作表名
        head: 工作表第一行
        """
        list_dir = file_name.split("\\")
        if not os.path.isfile(file_name):# 创建文件夹
            if len(list_dir) > 2:
                del list_dir[-1]
                newpath = "\\".join(list_dir)
                if not os.path.isdir(newpath):
                     os.makedirs(newpath)
            wb = openpyxl.Workbook()
            #
            if sheet_name not in wb:
                sh = wb.create_sheet(sheet_name)
            sh = wb[sheet_name]
            for nindex, data in enumerate(head):
                sh.cell(1, nindex + 1).value = str(data)  # 写文件
            wb.save(file_name)
            wb.close()


    def read_data_line(self):
        #按行读取数据转化为列表
        rows_data = list(self.sh.rows)
        if rows_data == []:
            return []
        list_list_data = []
        for currows_data in rows_data:
            list_data = []
            for data in currows_data:
                list_data.append(data.value)
            list_list_data.append(list_data)
        return list_list_data



    def writeExcel(self, list_list_data):
        """
        写入excel，叠加写入
        """
        self.b_iswriting = True
        ncurlinenum = self.sh.max_row
        for row in range(0,len(list_list_data)):
            list_data = list_list_data[row]
            for col in range(0,len(list_data)):
                self.sh.cell(row + 1 + ncurlinenum, col + 1).value = str(list_data[col])  # 写文件
        self.wb.save(self.file_name)
        self.b_iswriting = False


    def writeExcelinOverwrite(self, list_list_data, sheetname):
        """覆盖式写入, 且sheetname不存在会创建"""
        if sheetname not in self.wb:
            sh = self.wb.create_sheet(sheetname)
        else:
            self.clearsheet(sheetname)
            sh = self.wb[sheetname]
        self.b_iswriting = True
        for row in range(0,len(list_list_data)):
            list_data = list_list_data[row]
            for col in range(0,len(list_data)):
                sh.cell(row + 1, col + 1).value = str(list_data[col])  # 写文件
        self.wb.save(self.file_name)
        self.b_iswriting = False


    def read_data_line_for_sheet(self, sheet_name):
        #根据表格， 按行读取数据转化为列表
        if sheet_name in self.wb:
            sh = self.wb[sheet_name]
            rows_data = list(sh.rows)
            if rows_data == []:
                return []
            list_list_data = []
            for currows_data in rows_data:
                list_data = []
                for data in currows_data:
                    list_data.append(data.value)
                list_list_data.append(list_data)
            return list_list_data
        else:
            return []

    def clearsheet(self, sheetname):
        if sheetname in self.wb:
            rows_data = list(self.wb[sheetname].rows)
            if rows_data == []:
                return
            for row, currowsdata in enumerate(rows_data):
                self.wb[sheetname].delete_rows(1)
            # for row, currows_data in enumerate(rows_data):
            #     for col, data in enumerate(currows_data):
            #         self.wb[sheetname].cell(row + 1, col + 1).value = ''


if __name__ == '__main__':
    r = CExcelManager("D:\\MESLOG\\出站校验接口(dataCollectForSfcEx)\\2022-02-06.xlsx",'Sheet', ['条码', '开始时间', '结束时间',
                                                                                           '耗时', '传参', 'Code', 'message', '出站模式'])
    # list_list_data = [[1, 2, 3, 4], ['测试', '答复', 'ID', 6]]
    # r.writeExcel(list_list_data)
    # r.writeExcel(list_list_data)
    print(r.read_data_line())

