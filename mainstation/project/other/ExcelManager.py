import numpy as np
import cv2
import os

import openpyxl


class CExcelManager(object):
    """
    读写EXCEL的类
    """
    def __init__(self,file_name, sheet_name, head):
        """
        这个是用来初始化读取对象的
        :param file_name: 文件名 ---> str类型
        :param sheet_name: 表单名 ———> str类型
        """
        # 打开文件
        self.__createfile(file_name, sheet_name, head)
        self.wb = openpyxl.load_workbook(file_name)
        self.file_name = file_name
        # 选择表单
        self.sh = self.wb[sheet_name]
        self.b_iswriting = False# 是否正在写的状态


    def __createfile(self, file_name, sheet_name, head):
        """
        创建文件，包含创建文件夹
        file_name: 文件名
        sheet_name: 工作表名
        head: 工作表第一行
        """
        list_dir = file_name.split("\\")
        if not os.path.isfile(file_name):# 创建文件夹
            if len(list_dir) > 2:
                del list_dir[-1]
                newpath = "\\".join(list_dir)
                if not os.path.isdir(newpath):
                     os.makedirs(newpath)
            wb = openpyxl.Workbook()
            #
            if sheet_name not in wb:
                sh = wb.create_sheet(sheet_name)
            sh = wb[sheet_name]
            for nindex, data in enumerate(head):
                sh.cell(1, nindex + 1).value = str(data)  # 写文件
            wb.save(file_name)
            wb.close()


    def read_data_line(self):
        #按行读取数据转化为列表
        print(self.sh.max_row, self.sh.max_column)
        rows_data = list(self.sh.rows)
        if rows_data == []:
            return
        print (rows_data)
        # print(rows_data)
        # 获取表单的表头信息
        titles = []
        for title in rows_data[0]:
            titles.append(title.value)
        # print(titles)
        #定义一个空列表用来存储测试用例
        cases = []
        for case in rows_data[1:]:
            # print(case)
            data = []
            for cell in case: #获取一条测试用例数据
                # print(cell.value)
                data.append(cell.value)
                # print(data)
                #判断该单元格是否为字符串，如果是字符串类型则需要使用eval();如果不是字符串类型则不需要使用eval()
                if isinstance(cell.value,str):
                    data.append(eval(cell.value))
                else:
                    data.append(cell.value)
                #将该条数据存放至cases中
            # print(dict(list(zip(titles,data))))
                case_data = dict(list(zip(titles,data)))
                cases.append(case_data)
        return cases


    def writeExcel(self, list_list_data):
        self.b_iswriting = True
        ncurlinenum = self.sh.max_row
        for row in range(0,len(list_list_data)):
            list_data = list_list_data[row]
            for col in range(0,len(list_data)):
                self.sh.cell(row + 1 + ncurlinenum, col + 1).value = str(list_data[col])  # 写文件
        self.wb.save(self.file_name)
        self.b_iswriting = False


if __name__ == '__main__':
    r = CExcelManager("D:\\MESLOG\\出站校验接口(dataCollectForSfcEx)\\2022-02-06.xlsx",'Sheet', ['条码', '开始时间', '结束时间',
                                                                                           '耗时', '传参', 'Code', 'message', '出站模式'])
    list_list_data = [[1, 2, 3, 4], ['测试', '答复', 'ID', 6]]
    r.writeExcel(list_list_data)
    r.writeExcel(list_list_data)

